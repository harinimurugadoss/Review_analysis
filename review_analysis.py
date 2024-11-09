import pandas as pd
from pymongo import MongoClient
import openpyxl
from openpyxl.chart import BarChart, Reference
import sys

# Configuration - MODIFY THESE VALUES
CSV_FILE_PATH = r"C:\Users\ASUS\OneDrive\Desktop\last\amazon_review_scraper\configure\Samsung_Adaptor_reviews.csv"  # Adjust path if needed
RATING_COLUMN_NAME = "Rating"
COMMENT_COLUMN_NAME = "Comment"

# Categories for negative reviews
REVIEW_CATEGORIES = {
    'product_quality': ['defective', 'broken', 'poor quality', 'damaged', 'bad'],
    'customer_service': ['service', 'support', 'staff', 'representative', 'rude'],
    'delivery': ['shipping', 'delivery', 'late', 'delayed', 'packaging'],
    'price': ['expensive', 'overpriced', 'cost', 'price', 'waste of money'],
    'functionality': ["doesn't work", 'stopped working', 'not working', 'failed', 'issues', 'error']
}

def analyze_reviews(csv_file):
    try:
        print(f"Reading CSV file: {csv_file}")
        df = pd.read_csv(csv_file)

        if RATING_COLUMN_NAME not in df.columns:
            raise ValueError(f"Rating column '{RATING_COLUMN_NAME}' not found in CSV file")
        if COMMENT_COLUMN_NAME not in df.columns:
            raise ValueError(f"Comment column '{COMMENT_COLUMN_NAME}' not found in CSV file")

        # Create ratings distribution
        ratings_dist = df[RATING_COLUMN_NAME].value_counts().sort_index()

        # Create Excel workbook for chart
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ratings Distribution"

        # Add headers and data
        ws['A1'] = 'Rating'
        ws['B1'] = 'Count'

        for idx, (rating, count) in enumerate(ratings_dist.items(), start=2):
            ws[f'A{idx}'] = rating
            ws[f'B{idx}'] = count

        # Create and customize the bar chart
        chart = BarChart()
        chart.title = "Review Ratings Distribution"
        chart.x_axis.title = "Rating"
        chart.y_axis.title = "Number of Reviews"

        data = Reference(ws, min_col=2, min_row=1, max_row=len(ratings_dist) + 1, max_col=2)
        categories = Reference(ws, min_col=1, min_row=2, max_row=len(ratings_dist) + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        ws.add_chart(chart, "D2")

        # Save the Excel file
        output_file = "ratings_distribution.xlsx"
        wb.save(output_file)
        print(f"Chart saved to: {output_file}")

        # Return negative reviews (rating 1 or 2)
        return df[df[RATING_COLUMN_NAME].isin([1, 2])]

    except Exception as e:
        print(f"Error in analyze_reviews: {str(e)}")
        sys.exit(1)

def categorize_negative_reviews(negative_reviews):
    try:
        print("Connecting to MongoDB...")

        # Use your MongoDB connection string
        client = MongoClient(
            'mongodb+srv://harinimurugadoss4:Meeramurugadoss@cluster04.zlgp4.mongodb.net/?retryWrites=true&w=majority&appName=Cluster04',
            serverSelectionTimeoutMS=50000  # 50 seconds
        )

        db = client['review_analysis']
        collection = db['negative_reviews']

        # Clear existing data
        collection.delete_many({})
        print("Processing negative reviews...")

        for _, review in negative_reviews.iterrows():
            review_categories = []
            review_text = str(review[COMMENT_COLUMN_NAME]).lower()

            # Categorize reviews based on keywords
            for category, keywords in REVIEW_CATEGORIES.items():
                if any(keyword in review_text for keyword in keywords):
                    review_categories.append(category)

            if not review_categories:
                review_categories = ['other']

            document = {
                'rating': int(review[RATING_COLUMN_NAME]),
                'comment': review[COMMENT_COLUMN_NAME],
                'categories': review_categories,
                'processed_date': pd.Timestamp.now()
            }

            collection.insert_one(document)

        print("Negative reviews stored successfully!")
        return collection

    except Exception as e:
        print(f"Error in categorize_negative_reviews: {str(e)}")
        sys.exit(1)

def main():
    try:
        print("Starting review analysis...")

        # Create bar chart and get negative reviews
        negative_reviews = analyze_reviews(CSV_FILE_PATH)
        print(f"Found {len(negative_reviews)} negative reviews")

        # Categorize and store in MongoDB
        collection = categorize_negative_reviews(negative_reviews)

        # Print summary
        categories_summary = {}
        for doc in collection.find():
            for category in doc['categories']:
                categories_summary[category] = categories_summary.get(category, 0) + 1

        print("\nNegative Review Categories Summary:")
        for category, count in sorted(categories_summary.items()):
            print(f"{category}: {count} reviews")

        print("\nAnalysis completed successfully!")

    except Exception as e:
        print(f"Error in main: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
